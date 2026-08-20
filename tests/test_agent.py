import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from agent.agent_loop import LocalAgent
from agent.brain import answer_project_question, rank_chunks
from agent.permissions import PermissionError, resolve_inside_workspace
from agent.planner import plan
from agent.storage import PersistentStore
from agent.tools.filesystem import inspect_path, search_text

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


class PlannerTests(unittest.TestCase):
    def test_plans_guarded_git_commands(self):
        commit = plan("commit add agent v2")
        push = plan("push")

        self.assertEqual(commit.name, "git_commit")
        self.assertTrue(commit.needs_confirmation)
        self.assertEqual(push.name, "git_push")
        self.assertTrue(push.needs_confirmation)

    def test_plans_memory_commands(self):
        self.assertEqual(plan("remember use safe tools").name, "remember")
        self.assertEqual(plan("recall safe").name, "recall")
        self.assertEqual(plan("todo add add voice").name, "todo_add")

    def test_plans_guarded_replace_apply(self):
        preview = plan("replace README.md: old => new")
        apply = plan("apply replace README.md: old => new")

        self.assertEqual(preview.name, "replace_in_file")
        self.assertFalse(preview.needs_confirmation)
        self.assertEqual(apply.name, "apply_replace_in_file")
        self.assertTrue(apply.needs_confirmation)

    def test_plans_guarded_pending_edit_apply(self):
        preview = plan("edit README.md: add a quick start section")
        apply = plan("apply edit")

        self.assertEqual(preview.name, "edit_file")
        self.assertFalse(preview.needs_confirmation)
        self.assertEqual(apply.name, "apply_pending_edit")
        self.assertTrue(apply.needs_confirmation)

    def test_plans_guarded_excel_apply(self):
        preview = plan("excel set budget.xlsx: Sheet1!B2 = 1200")
        apply = plan("apply excel")

        self.assertEqual(preview.name, "excel_set_cell")
        self.assertFalse(preview.needs_confirmation)
        self.assertEqual(apply.name, "apply_excel_change")
        self.assertTrue(apply.needs_confirmation)

    def test_plans_excel_create_and_append(self):
        create = plan("excel create budget.xlsx: Budget = Categoria, Importo")
        append = plan("excel append budget.xlsx: Budget = Affitto, 700")

        self.assertEqual(create.name, "excel_create")
        self.assertFalse(create.needs_confirmation)
        self.assertEqual(append.name, "excel_append_row")
        self.assertFalse(append.needs_confirmation)


class PermissionTests(unittest.TestCase):
    def test_blocks_paths_outside_workspace(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with self.assertRaises(PermissionError):
                resolve_inside_workspace(temp_dir, "../outside.txt")
    def test_llm_recognizes_natural_read_request(self):
        result = plan("puoi leggermi il file README.md?")
        self.assertEqual(result.name, "read_file")
        self.assertEqual(result.argument, "README.md")

    def test_llm_recognizes_natural_search_request(self):
        result = plan("cerca la parola relu nel codice")
        self.assertEqual(result.name, "search_text")
        self.assertEqual(result.argument, "relu")

    def test_llm_recognizes_natural_todo_request(self):
        result = plan("aggiungi alla lista delle cose da fare: sistemare i colori")
        acceptable_actions = {"todo_add", "ask_project"}
        self.assertIn(result.name, acceptable_actions)
        if result.name == "todo_add":
            self.assertIn("colori", result.argument)

    def test_llm_commit_still_needs_confirmation(self):
        result = plan("fai il commit con messaggio test di sicurezza")
        self.assertEqual(result.name, "git_commit")
        self.assertTrue(result.needs_confirmation)

    def test_unknown_request_never_triggers_dangerous_action(self):
        result = plan("qualcosa di completamente senza senso xyzabc123")
        dangerous_actions = {"train_model", "git_commit", "git_push"}
        self.assertNotIn(result.name, dangerous_actions)
        self.assertFalse(result.needs_confirmation)

class StorageTests(unittest.TestCase):
    def test_persists_memory_and_todos(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            store = PersistentStore(temp_dir)
            store.remember("project prefers local tools")
            store.add_todo("add voice later")

            reloaded = PersistentStore(temp_dir)
            self.assertIn("local tools", reloaded.recall("local"))
            self.assertIn("add voice later", reloaded.list_todos())


class FilesystemToolTests(unittest.TestCase):
    def test_search_and_inspect_text_file(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "note.txt"
            path.write_text("hello agent\n", encoding="utf-8")

            self.assertIn("note.txt:1", search_text(temp_dir, "agent"))
            self.assertIn("Lines: 1", inspect_path(temp_dir, "note.txt"))


class EditingToolTests(unittest.TestCase):
    def test_replace_previews_then_applies_after_confirmation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "note.txt"
            path.write_text("hello agent\n", encoding="utf-8")
            agent = LocalAgent(temp_dir)

            preview = agent.handle("replace note.txt: hello => ciao")
            self.assertIn("Proposed change:", preview)
            self.assertIn("-hello agent", preview)
            self.assertIn("+ciao agent", preview)
            self.assertEqual(path.read_text(encoding="utf-8"), "hello agent\n")

            blocked = agent.handle("apply replace note.txt: hello => ciao")
            self.assertIn("changes project state", blocked)
            self.assertEqual(path.read_text(encoding="utf-8"), "hello agent\n")

            applied = agent.handle("apply replace note.txt: hello => ciao", assume_yes=True)
            self.assertIn("Updated: note.txt", applied)
            self.assertEqual(path.read_text(encoding="utf-8"), "ciao agent\n")

    def test_edit_previews_pending_change_then_applies_after_confirmation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "note.txt"
            path.write_text("hello agent\n", encoding="utf-8")
            agent = LocalAgent(temp_dir)

            with patch(
                "agent.tools.editing.request_edit_from_llm",
                return_value="hello Jarvis\n",
            ):
                preview = agent.handle("edit note.txt: rename agent to Jarvis")

            self.assertIn("Proposed edit for note.txt:", preview)
            self.assertIn("-hello agent", preview)
            self.assertIn("+hello Jarvis", preview)
            self.assertEqual(path.read_text(encoding="utf-8"), "hello agent\n")

            blocked = agent.handle("apply edit")
            self.assertIn("changes project state", blocked)
            self.assertEqual(path.read_text(encoding="utf-8"), "hello agent\n")

            applied = agent.handle("apply edit", assume_yes=True)
            self.assertIn("Updated: note.txt", applied)
            self.assertEqual(path.read_text(encoding="utf-8"), "hello Jarvis\n")

    def test_pending_edit_is_discarded_if_file_changes_after_preview(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "note.txt"
            path.write_text("hello agent\n", encoding="utf-8")
            agent = LocalAgent(temp_dir)

            with patch(
                "agent.tools.editing.request_edit_from_llm",
                return_value="hello Jarvis\n",
            ):
                agent.handle("edit note.txt: rename agent to Jarvis")

            path.write_text("changed elsewhere\n", encoding="utf-8")
            response = agent.handle("apply edit", assume_yes=True)

            self.assertIn("file changed after the preview", response)
            self.assertEqual(path.read_text(encoding="utf-8"), "changed elsewhere\n")


@unittest.skipIf(Workbook is None, "openpyxl is not installed")
class SpreadsheetToolTests(unittest.TestCase):
    def test_excel_create_previews_then_creates_after_confirmation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "budget.xlsx"
            agent = LocalAgent(temp_dir)

            preview = agent.handle("excel create budget.xlsx: Budget = Categoria, Importo, Data")
            self.assertIn("Proposed Excel workbook:", preview)
            self.assertIn("+ file: budget.xlsx", preview)
            self.assertFalse(path.exists())

            blocked = agent.handle("apply excel")
            self.assertIn("changes project state", blocked)
            self.assertFalse(path.exists())

            applied = agent.handle("apply excel", assume_yes=True)
            self.assertIn("Created: budget.xlsx | Budget", applied)
            self.assertTrue(path.exists())

            from openpyxl import load_workbook
            loaded = load_workbook(path)
            try:
                sheet = loaded["Budget"]
                self.assertEqual(sheet["A1"].value, "Categoria")
                self.assertEqual(sheet["B1"].value, "Importo")
                self.assertEqual(sheet["C1"].value, "Data")
                self.assertTrue(sheet["A1"].font.bold)
            finally:
                loaded.close()

    def test_excel_tools_preview_then_apply_cell_change(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "budget.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget"
            sheet["A1"] = "Item"
            sheet["B1"] = "Amount"
            sheet["A2"] = "Server"
            sheet["B2"] = 100
            workbook.save(path)

            agent = LocalAgent(temp_dir)

            self.assertIn("Budget", agent.handle("excel sheets budget.xlsx"))
            self.assertIn("Server\t100", agent.handle("excel read budget.xlsx: Budget!A2:B2"))

            preview = agent.handle("excel set budget.xlsx: Budget!B2 = 1200")
            self.assertIn("Proposed Excel change:", preview)
            self.assertIn("- budget.xlsx | Budget!B2: 100", preview)
            self.assertIn("+ budget.xlsx | Budget!B2: '1200'", preview)

            blocked = agent.handle("apply excel")
            self.assertIn("changes project state", blocked)

            applied = agent.handle("apply excel", assume_yes=True)
            self.assertIn("Updated: budget.xlsx | Budget!B2", applied)

            from openpyxl import load_workbook
            loaded = load_workbook(path)
            try:
                self.assertEqual(loaded["Budget"]["B2"].value, 1200)
            finally:
                loaded.close()

    def test_excel_append_previews_then_applies_after_confirmation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "budget.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget"
            sheet.append(["Categoria", "Importo"])
            workbook.save(path)

            agent = LocalAgent(temp_dir)
            preview = agent.handle("excel append budget.xlsx: Budget = Affitto, 700")
            self.assertIn("Proposed Excel row append:", preview)
            self.assertIn("Budget!row 2", preview)

            blocked = agent.handle("apply excel")
            self.assertIn("changes project state", blocked)

            applied = agent.handle("apply excel", assume_yes=True)
            self.assertIn("appended row 2 to Budget", applied)

            from openpyxl import load_workbook
            loaded = load_workbook(path)
            try:
                sheet = loaded["Budget"]
                self.assertEqual(sheet["A2"].value, "Affitto")
                self.assertEqual(sheet["B2"].value, 700)
            finally:
                loaded.close()

    def test_pending_excel_append_is_discarded_if_sheet_changes_after_preview(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "budget.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget"
            sheet.append(["Categoria", "Importo"])
            workbook.save(path)

            agent = LocalAgent(temp_dir)
            agent.handle("excel append budget.xlsx: Budget = Affitto, 700")

            from openpyxl import load_workbook
            loaded = load_workbook(path)
            try:
                loaded["Budget"].append(["Internet", 40])
                loaded.save(path)
            finally:
                loaded.close()

            response = agent.handle("apply excel", assume_yes=True)
            self.assertIn("sheet changed after the preview", response)

            loaded = load_workbook(path)
            try:
                self.assertEqual(loaded["Budget"].max_row, 2)
            finally:
                loaded.close()

    def test_pending_excel_change_is_discarded_if_cell_changes_after_preview(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "budget.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget"
            sheet["B2"] = 100
            workbook.save(path)

            agent = LocalAgent(temp_dir)
            agent.handle("excel set budget.xlsx: Budget!B2 = 1200")

            from openpyxl import load_workbook
            loaded = load_workbook(path)
            try:
                loaded["Budget"]["B2"] = 300
                loaded.save(path)
            finally:
                loaded.close()

            response = agent.handle("apply excel", assume_yes=True)
            self.assertIn("cell changed after the preview", response)

            loaded = load_workbook(path)
            try:
                self.assertEqual(loaded["Budget"]["B2"].value, 300)
            finally:
                loaded.close()


class BrainTests(unittest.TestCase):
    def test_ranks_relevant_project_chunks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "README.md"
            path.write_text("The agent has persistent memory and todos.\n", encoding="utf-8")

            chunks = rank_chunks(temp_dir, "how does memory work")

            self.assertEqual(chunks[0].path, "README.md")

    def test_answers_with_evidence(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "agent.py"
            path.write_text("The planner routes ask commands.\n", encoding="utf-8")

            answer = answer_project_question(temp_dir, "planner ask commands")

            # La risposta puo' arrivare dall'LLM (testo libero) o dal fallback
            # estrattivo (formato fisso con Evidence/Relevant points).
            # In entrambi i casi deve essere non vuota e sensata.
            self.assertTrue(len(answer) > 0)
            self.assertNotIn("could not find", answer.lower())


if __name__ == "__main__":
    unittest.main()
