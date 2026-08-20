from agent.permissions import resolve_inside_workspace


def require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils.cell import range_boundaries
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is not installed. Run `pip install -r requirements.txt`."
        ) from error

    return Workbook, load_workbook, Font, range_boundaries


def resolve_workbook_path(workspace_root, user_path):
    target = resolve_inside_workspace(workspace_root, user_path)
    if target.suffix.lower() != ".xlsx":
        raise ValueError("Excel tools only support .xlsx files.")
    return target


def list_sheets(workspace_root, user_path):
    _, load_workbook, _, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)

    workbook = load_workbook(target, read_only=True, data_only=False)
    try:
        return "\n".join(workbook.sheetnames)
    finally:
        workbook.close()


def read_range(workspace_root, user_path, sheet_name, range_ref):
    _, load_workbook, _, range_boundaries = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)

    workbook = load_workbook(target, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        rows = []
        for row in sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=False,
        ):
            values = []
            for cell in row:
                value = cell.value
                values.append("" if value is None else str(value))
            rows.append("\t".join(values))

        return "\n".join(rows) if rows else "(empty range)"
    finally:
        workbook.close()


def get_cell_value(workspace_root, user_path, sheet_name, cell_ref):
    _, load_workbook, _, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)

    workbook = load_workbook(target, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return workbook[sheet_name][cell_ref].value
    finally:
        workbook.close()


def preview_set_cell(workspace_root, user_path, sheet_name, cell_ref, value):
    current_value = get_cell_value(workspace_root, user_path, sheet_name, cell_ref)
    return (
        f"Proposed Excel change:\n"
        f"- {user_path} | {sheet_name}!{cell_ref}: {format_cell_value(current_value)}\n"
        f"+ {user_path} | {sheet_name}!{cell_ref}: {format_cell_value(value)}"
    )


def apply_set_cell(workspace_root, user_path, sheet_name, cell_ref, value):
    _, load_workbook, _, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)

    workbook = load_workbook(target)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        cell = sheet[cell_ref]
        cell.value = coerce_cell_value(value)

        temp_path = target.with_name(target.name + ".tmp.xlsx")
        workbook.save(temp_path)
    finally:
        workbook.close()

    temp_path.replace(target)
    return f"Updated: {user_path} | {sheet_name}!{cell_ref}"


def preview_create_workbook(workspace_root, user_path, sheet_name, headers):
    target = resolve_workbook_path(workspace_root, user_path)
    if target.exists():
        raise FileExistsError(user_path)
    if not sheet_name:
        raise ValueError("Sheet name cannot be empty.")
    if not headers:
        raise ValueError("At least one column is required.")

    return (
        "Proposed Excel workbook:\n"
        f"+ file: {user_path}\n"
        f"+ sheet: {sheet_name}\n"
        f"+ headers: {', '.join(headers)}"
    )


def apply_create_workbook(workspace_root, user_path, sheet_name, headers):
    Workbook, _, Font, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if target.exists():
        raise FileExistsError(user_path)
    if not sheet_name:
        raise ValueError("Sheet name cannot be empty.")
    if not headers:
        raise ValueError("At least one column is required.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for column_index, header in enumerate(headers, start=1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = max(12, min(len(str(header)) + 4, 32))

    target.parent.mkdir(parents=True, exist_ok=True)
    temp_path = target.with_name(target.name + ".tmp.xlsx")
    workbook.save(temp_path)
    workbook.close()

    temp_path.replace(target)
    return f"Created: {user_path} | {sheet_name}"


def get_sheet_max_row(workspace_root, user_path, sheet_name):
    _, load_workbook, _, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)

    workbook = load_workbook(target, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        return workbook[sheet_name].max_row
    finally:
        workbook.close()


def preview_append_row(workspace_root, user_path, sheet_name, values):
    next_row = get_sheet_max_row(workspace_root, user_path, sheet_name) + 1
    if not values:
        raise ValueError("At least one value is required.")

    return (
        "Proposed Excel row append:\n"
        f"+ {user_path} | {sheet_name}!row {next_row}: "
        f"{', '.join(format_cell_value(value) for value in values)}"
    )


def apply_append_row(workspace_root, user_path, sheet_name, values):
    _, load_workbook, _, _ = require_openpyxl()
    target = resolve_workbook_path(workspace_root, user_path)

    if not target.exists():
        raise FileNotFoundError(user_path)
    if not values:
        raise ValueError("At least one value is required.")

    workbook = load_workbook(target)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")

        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1
        sheet.append([coerce_cell_value(value) for value in values])

        temp_path = target.with_name(target.name + ".tmp.xlsx")
        workbook.save(temp_path)
    finally:
        workbook.close()

    temp_path.replace(target)
    return f"Updated: {user_path} | appended row {next_row} to {sheet_name}"


def parse_excel_location(location):
    if "!" not in location:
        return None

    sheet_name, cell_ref = location.split("!", 1)
    sheet_name = sheet_name.strip().strip("'").strip('"')
    cell_ref = cell_ref.strip().upper()

    if not sheet_name or not cell_ref:
        return None

    return sheet_name, cell_ref


def parse_comma_values(text):
    return [value.strip() for value in text.split(",") if value.strip()]


def coerce_cell_value(value):
    text = value.strip()
    if text.startswith("="):
        return text
    if text.lower() in {"true", "false"}:
        return text.lower() == "true"

    try:
        if "." not in text and "," not in text:
            return int(text)
    except ValueError:
        pass

    try:
        normalized = text.replace(",", ".")
        return float(normalized)
    except ValueError:
        return value


def format_cell_value(value):
    if value is None:
        return "(blank)"
    return repr(value)
